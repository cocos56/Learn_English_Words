import openpyxl
import xlsxwriter
import re

def removeBeginAndEndSpaces(s):
    if s == None:
        return None
    i = 0
    l = len(s)
    while(i<l and s[i]==' '):
        i += 1
    s = s[i:]
    l = len(s)
    if(l == 0):
        return s
    i = len(s)-1
    while(s[i]==' '):
        i -= 1
    s = s[:i+1]
    return s

Wli = [
    'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'
    , 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z'
    ,' ', '-', '.' ,'\'' , 'é'
    ]

def judge(s):
    for i in s:
        if i not in Wli:
            return False
    return True

List = []

for i in '  ':
    print(ord(i))

fileName = 'WordsList_CET4.xlsx'
#fileName = 'WordsList_CET6.xlsx'
#fileName = 'WordsList_考研5500.xlsx'
#fileName = 'WorksList_高考3500.xlsx'

wb = openpyxl.load_workbook(fileName)
workbook = xlsxwriter.Workbook('test2.xlsx')
worksheet = workbook.add_worksheet()
print(wb.sheetnames)
count = 1
ws = wb[wb.sheetnames[0]]
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    li = range(1, ws.max_row+1)
    for i in li:
        for c in range(1,2):
            v = ws.cell(i,c)
            if not v in List:
                List.append(v)
            else:
                print(i, v)


wb.save(fileName)