import os
import xlrd
import xlsxwriter
import FN
import re
import json
from openpyxl import load_workbook

class Unipus:
    def __init__(self):
        self.dict = {
            3 : [1, 2, 4, 5, 8]
            }
        self.wordsType = ['vi.', 'vt.', 'a.', 'n.', 'ad.', 'v.', 'prep.', 'suffix', 'infml.']
        self.wordsType2 = []
        self.d = os.getcwd() + '\\Datas'
        self.fn = FN.FN(d=self.d + '\\RawXlsx')
        self.fn.analyzeExtensions()
        self.filesName = []
        for i in self.fn.filesName:
            self.filesName.append(self.fn.analyzeFN(i))
            pass
    
    def seekWT(self):
        for i in self.filesName:
            self.__seekWT(i)
        pass

    def __seekWT(self, fileName):
        data_file = fileName[0] + fileName[1] + fileName[2]
        print(data_file)
        workbook = xlrd.open_workbook(data_file)
        sheet1 = workbook.sheet_by_index(0)
        if (not (sheet1.nrows and sheet1.ncols)):
            return 

        begin = 0
        while(sheet1.cell(begin,0).value == '??'):
            begin += 1
        while(sheet1.cell(begin,0).value == chr(65279)):
            begin += 1
        for i in range(begin,sheet1.nrows):
            data = sheet1.cell(i, 0).value
            if not data=='':
                data = self.__remove_end_space(data)
                if((self.__check_contain_chinese(data) or not data.find('(')== -1)):
                    res = re.findall('.*?([a-z.]+).*?' ,data)
                    if(not res == []):
                        if(len(res)==1):
                            if not res[0] in self.wordsType and not res[0] in self.wordsType2:
                                print(data)
                                self.wordsType2.append(res[0])
                        elif not res[0].find(r'.') == -1:
                            print(data)
                            print(res)
        pass

    def modifyXlsx(self):
        for i in self.filesName:
            self.__modifyXlsx(i)
        pass

    def __modifyXlsx(self, fileName):
        data_file = fileName[0] + fileName[1] + fileName[2]
        workbook = xlrd.open_workbook(data_file)
        sheetnames = workbook.sheet_names()
        d_f = self.d + '\\ModifyXlsx\\'  + fileName[1] + fileName[2]
        wb = xlsxwriter.Workbook(d_f)
        cell_format_center = wb.add_format({'align':'center', 'valign':'vcenter'})
        title = ['单词', '音标', '用户释义','词典释义', '考研释义', '六级释义', '四级释义', '高考释义', '只记词汇']
        width = []
        for i in title:
            width.append(len(i))
        for sheetN in sheetnames:
            sheet = workbook.sheet_by_name(sheetN)
            ws = wb.add_worksheet(sheetN)
            for i in range(len(title)):
                ws.write(0, i ,title[i], cell_format_center)
                k = 0
                begin = 0
                while(sheet.cell(begin,0).value == '??'):
                    begin += 1
                while(sheet.cell(begin,0).value == chr(65279)):
                    begin += 1
                for i in range(begin,sheet.nrows):
                    data = sheet.cell(i, 0).value
                    if not data=='':
                        data = self.__remove_end_space(data)
                        if(not self.__check_contain_chinese(data)):
                            if(not data.find('(')== -1 and not data.find(',')==-1):
                                flag = False
                            else:
                                flag = True
                        else:
                            flag = False
                        if(flag):
                            k += 1
                            l = self.__caclutWidth(data)
                            if(l>width[0]):
                                width[0] = l
                            ws.write(k, 0, data)
                            f = True
                        else:
                            if(f):
                                l = self.__caclutWidth(data)
                                if(l>width[2]):
                                    width[2] = l
                                ws.write(k, 2, data)
                                temp = data
                                f = False
                            else:
                                l = self.__caclutWidth(data)
                                if(l>width[2]):
                                    width[2] = l
                                temp =  temp + '\n' + data
                                ws.write(k, 2, temp)
                ws.set_column(0, 0, width = width[0])
                ws.set_column(2, 2, width = width[2])
                print(width[0])
        wb.close()

    def __caclutWidth(self, s):
        l = len(s)
        if(l>=56):
            l = 56
        return l
    
    def __check_contain_chinese(self, check_str):
        for ch in check_str:
            if u'\u4e00' <= ch <= u'\u9fff':
                return True
        return False
    
    def __modifySpaceAndQM(self,s):
        '''
        空格有32的空格和160的空格，32的空格为常见空格，
        这里将160的空格统一转成32的
        '''
        s = s.replace(chr(160), chr(32))
        s = s.replace('‘', '\'')
        s = s.replace('’', '\'')
        return s

    def __remove_end_space(self, s):
        s = self.__modifySpaceAndQM(s)
        length = len(s)
        while(length and s[length-1] == ' '):
            length -= 1
        return s[:length]

class Xdf():
    def __init__(self):
        self.d = r'D:\0COCO\本科\英语四六级\六级词汇词根+联想记忆法(MP3+文本)'
        self.fn = FN.FN(d=self.d)
        self.fn.analyzeExtensions()
        self.filesName = []
        self.wordsType = ['vt.', 'adj.', 'n.', 'v.', 'vi.', 'adv.', 'prep.', 'conj.']
        self.wordsType2 = []
        for i in self.fn.filesName:
            temp = self.fn.analyzeFN(i)
            if temp[2] == '.lrc':
                self.filesName.append(temp)
            pass

    def TranformXlsxToJson(self):
        xlsxF = r'D:\0COCO\本科\英语四六级\六级词汇词根+联想记忆法(MP3+文本)\000wordList.xlsx'
        jsonF = r'D:\0COCO\本科\英语四六级\六级词汇词根+联想记忆法(MP3+文本)\000wordList.json'
        
        wb = xlrd.open_workbook(xlsxF)
        s1 = wb.sheet_by_index(0)
        workbook = load_workbook(xlsxF)
        worksheet = workbook['Sheet1']
        for i in range(1, s1.nrows+1):
            worksheet.cell(row=i, column=2).value = self.remove_begin_space(worksheet.cell(row=i, column=2).value)
        workbook.save(xlsxF)

        WDict = {}
        for i in range(s1.nrows):
            d = {s1.cell(i,0).value : s1.cell(i,1).value}
            WDict.update(d)
        with open(jsonF, 'w', encoding='utf-8') as f:
            json.dump(WDict,f, ensure_ascii=False)
            f.close()
        with open(jsonF, 'r', encoding='utf-8') as f:
            data = json.load(f)
            #print(data)
            f.close()

    def combineXlsx(self):
        filesName = []
        for i in self.fn.filesName:
            temp = self.fn.analyzeFN(i)
            if temp[2] == '.xlsx' and not temp[1] == '000wordList':
                filesName.append(temp)
        workbook = xlsxwriter.Workbook(filesName[0][0] + '000wordList.xlsx')
        worksheet = workbook.add_worksheet()
        j = 0
        for i in filesName:
            i = i[0] + i[1] + i[2]
            print(i)
            wb = xlrd.open_workbook(i)
            s1 = wb.sheet_by_index(0)
            print(s1.name, s1.nrows, s1.ncols)

            for i in range(s1.nrows):
                #print(s1.cell(i,0).value, s1.cell(i,1).value)
                if(not s1.cell(i,0).value==''):
                    worksheet.write(j, 0, s1.cell(i,0).value)
                    temp = s1.cell(i,1).value.replace('\n', '')
                    worksheet.write(j, 1, temp)
                    j += 1
                else:
                    temp = temp + ' ' + s1.cell(i,1).value
                    temp = temp.replace('\n', '')
                    worksheet.write(j-1, 1, temp)
        workbook.close()

    def remove_begin_space(self, s):
        l = len(s)
        i = 0
        while i<l:
            if s[i] == ' ':
                i += 1
            else:
                break
        return s[i:]

    def getDataFromTxt(self):
        for i in self.filesName:
            txtF = i[0] + i[1] + '.txt'
            if(os.path.exists(txtF)):
                xlsxF = i[0] + i[1] + '.xlsx'
                self.__getDataFromTxt(xlsxF, txtF)
        pass
    def __getDataFromTxt(self, filename, txtF):
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        i = 0
        with open(txtF, 'r', encoding='utf-8') as f:
            line = f.readline()
            while(line):
                index = line.find(' ')
                data = line[:index]
                data2 = line[index:]
                if(data.isalpha()):
                    worksheet.write(i, 0, data)
                else:
                    data2 = line
                worksheet.write(i, 1, data2)
                i += 1
                line = f.readline()
        workbook.close()
        pass

    def seekWT(self):
        for i in self.filesName:
            txtF = i[0] + i[1] + '.txt'
            if(os.path.exists(txtF)):
                self.__seekWT(txtF)
        pass

    def __seekWT(self, fileName):
        print(fileName)
        with open(fileName, 'r', encoding = 'utf-8') as f:
                line = f.readline()
                while(line):
                    index = line.find('. ')
                    if (not index==-1):
                        data = line[:index+1]
                        res = re.findall('.*?([a-z.]+).*?' ,data)
                        for i in res:
                            if(not i.find('.')==-1):
                                if not i in self.wordsType and not i in self.wordsType2:
                                    print(data)
                                    self.wordsType2.append(res[0])
                    line = f.readline()
                f.close()
        pass

    def getDataFromLrc(self):
        for i in self.filesName:
            lrcF = i[0] + i[1] + i[2]
            txtF = i[0] + i[1] + '.txt'
            text = []
            if True:
            #if not os.path.exists(txtF):
                with open(lrcF, 'r', encoding = 'utf-8') as f:
                    line = f.readline()
                    while(line):
                        if not line.find('Word List ') == -1:
                            line = f.readline()
                            break
                        line = f.readline()
                    while(line):
                        res = re.findall('\[\d{2}:\d{2}\.\d{2}\](.+)', line)
                        if(res):
                            print(res)
                            text.append(res[0])
                        line = f.readline()
                        pass
                    f.close()
                    pass
                with open(txtF, 'w', encoding = 'utf-8') as f:
                    for i in text:
                        i = i.replace(chr(9), chr(32))
                        f.write(i+'\n')
                    f.close()
                    pass
                #print(text)
        pass

class Support():
    def __init__(self):
        self.cwd = os.getcwd()
        pass
    
    def __getFilesName(self, dir, typeList):
        filesName = []
        fn = FN.FN(dir)
        fn.analyzeExtensions()
        for i in fn.filesName:
            temp = fn.analyzeFN(i)
            if temp[2] in typeList:
                filesName.append(temp)
        return filesName

    def __removeBeginAndEndSpaces(self, s):
        i = 0
        while(s[i]==' '):
            i += 1
        s = s[i:]
        i = len(s)-1
        while(s[i]==' '):
            i -= 1
        s = s[:i+1]
        return s

    def TranformXlsxToJson(self):
        WordsListDir = self.cwd + r'\Datas\WordList'
        TL = ['.xlsx']
        filesName = self.__getFilesName(WordsListDir, TL)
        for i in filesName:
            xlsxF = i[0] + i[1] + i[2]
            jsonF = i[0] + i[1] + '.json'
            self.__TranformXlsxToJson(xlsxF, jsonF)
        
    def __TranformXlsxToJson(self, xlsxF, jsonF):
        wb = load_workbook(xlsxF)
        for sheetName in wb.sheetnames:
            ws = wb[sheetName]
            li = range(1, ws.max_row+1)
            li2 = range(1,3)
            for r in li:
                for c in li2:
                    ws.cell(r, c).value = self.__removeBeginAndEndSpaces(ws.cell(r, c).value)
        wb.save(xlsxF)

        WDict = {}
        for sheetName in wb.sheetnames:
            ws = wb[sheetName]
            li = range(1, ws.max_row+1)
            li2 = range(1,3)
            for r in li:
                d = {ws.cell(r, 1).value : ws.cell(r, 2).value}
                WDict.update(d)
        with open(jsonF, 'w', encoding='utf-8') as f:
            json.dump(WDict,f, ensure_ascii=False)
            f.close()
        # with open(jsonF, 'r', encoding='utf-8') as f:
        #     data = json.load(f)
        #     #print(data)
        #     f.close()

if __name__ == '__main__':
    unipus = Unipus()
    unipus.modifyXlsx()
    # support = Support()
    # support.TranformXlsxToJson()
