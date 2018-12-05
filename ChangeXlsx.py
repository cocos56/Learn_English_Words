from openpyxl import load_workbook
import os
import xlrd
import Spyder
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import FN
import ModifyData
import wget
import json

class ChangeXlsx():
    def __init__(self, recreate = False):
        self.d = os.getcwd() + '\\Datas'
        self.fn = FN.FN(d=self.d + '\\ModifyXlsx')
        self.fn.analyzeExtensions()
        self.font = Font(name='Arial',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
        self.space = [' ', ' ']
        if(recreate):
            md = ModifyData.Unipus()
            md.modifyXlsx()
        self.spy = Spyder.Spy()
        if(not os.path.exists(self.d + '\\Mp3')):
            os.mkdir(self.d + '\\Mp3')
        self.wb = None
        pass
    
    def addData(self):
        for file_dir in self.fn.filesName:
            self.wb = load_workbook(filename= file_dir)
            self.addPhoneticSymbol()
            self.addMeaning()
            self.__addSpecialMeaning()
            self.__setStyle()
            self.wb.save(file_dir)
    
    def __setStyle(self):
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            for i in range(2, ws.max_row+1):
                for j in range(1, ws.max_column+1):
                    ws.cell(i,j).alignment = Alignment(wrap_text=True,
                                                        vertical='center')
    def __addSpecialMeaning(self):
        Files = ['考研5500', 'CET6', 'CET4', '高考3500']
        Dict = []
        for i in Files:
            i = self.d + '\\' + 'WordList\\' + i + '.json'
            with open(i,'r',encoding='utf-8') as f:
                Dict.append(json.load(f))
                f.close()
        print(Dict)
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            for c in range(4):
                width = self.__caclutWidth(ws.cell(1,5+c).value)
                for i in range(2, ws.max_row+1):
                    if(ws.cell(i,1).value in Dict[c]):
                        ws.cell(i,5+c).value = Dict[c][ws.cell(i,1).value]
                        w = self.__caclutWidth(ws.cell(i,5+c).value)
                        if w > width:
                            width = w
                ws.column_dimensions[chr(ord('E')+c)].width = width

    def addPhoneticSymbol(self):
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            width = self.__caclutWidth(ws.cell(1,2).value)
            for i in range(2, ws.max_row+1):
                if(self.__isWord(ws.cell(row=i, column=1).value)):
                    ws.cell(row=i, column=2).font = self.font
                    #print(ws.cell(row=i, column=1).value)
                    ps = self.spy.getData(ws.cell(row=i, column=1).value)
                    #print(ps)
                    if not ps == '':
                        ws.cell(row=i, column=2).value = ps[0][0]
                        w = self.__caclutWidth(ws.cell(i,2).value)
                        if w > width:
                            width = w
                        self.__getMp3(ws.cell(row=i, column=1).value, ps[0][1])
            ws.column_dimensions['B'].width = width

    def addMeaning(self):
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            width = self.__caclutWidth(ws.cell(1,4).value)
            for i in range(2, ws.max_row+1):
                if(self.__isWord(ws.cell(row=i, column=1).value)):
                    #print(ws.cell(row=i, column=1).value)
                    ps = self.spy.getMeaning(ws.cell(row=i, column=1).value)
                    str1 = ''
                    for i1 in ps:
                        str2 = i1[0] + '  '
                        str3 = ''
                        for i3 in i1[1]:
                            str3 += i3
                        str2 += str3
                        str1 += str2
                    #print(str1)
                    ws.cell(row=i, column=4).value = str1
                    w = self.__caclutWidth(ws.cell(i,4).value)
                    if w > width:
                        width = w
            ws.column_dimensions['D'].width = width

    def __caclutWidth(self, s):
        l = len(s)
        if(l>=56):
            l = 56
        return l

    def __getMp3(self, word, url):
        if(not os.path.exists(self.d + '\\Mp3\\' + word + '.mp3')):
            wget.download(url, out=self.d + '\\Mp3\\' + word + '.mp3')
        pass


    def __isWord(self, str):
        if str == None:
            return False
        if not str == '':
            return  str.find(' ') == -1

if __name__ == '__main__':
    recreate = True
    ch = ChangeXlsx(recreate=recreate)
    ch.addData()
    #ch.addSpecialMeaning()
    pass